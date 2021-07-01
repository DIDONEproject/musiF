import copy
from musif.extract.features import key
from musif.config import Configuration
import os
import sys
import warnings
from itertools import combinations, permutations
from multiprocessing import Lock
from os import path
from typing import Dict, List

import musif.extract.features.ambitus as ambitus
import musif.extract.features.interval as interval
import musif.extract.features.lyrics as lyrics
import numpy as np
import openpyxl
import pandas as pd
from matplotlib.pyplot import axis
from musif.common.sort import sort, sort_dataframe
from openpyxl.styles.fonts import Font
from openpyxl.writer.excel import ExcelWriter
from pandas.core.frame import DataFrame
from tqdm import tqdm

from .harmony_sorting import * #TODO: REVIEW

from .calculations import (compute_average, compute_value,
                           make_intervals_absolute)
from .constants import *
from .utils import (adjust_excel_width_height, columns_alike_our_data,
                    get_groups_add_info,
                    prepare_data_emphasised_scale_degrees_second,
                    print_averages_total, print_averages_total_column,
                    remove_folder_contents, split_voices, write_columns_titles,
                    write_columns_titles_variable_length)
from .visualisations import (bar_plot, bar_plot_extended, box_plot,
                             customized_plot, double_bar_plot,
                             line_plot_extended, melody_bar_plot, pie_plot)
#Initialize global config vaiable
cgf=None
rows_groups={}
not_used_cols={}

if not os.path.exists(path.join(os.getcwd(), 'logs')):
    os.mkdir(path.join(os.getcwd(), 'logs'))


#####################################################################
def _tasks_execution(rg: dict, nuc:list, _cfg: Configuration, groups: list, results_path_factorx: str, additional_info, factor: int, common_columns_df: DataFrame, 
notes_df: DataFrame = pd.DataFrame(), melody_values: DataFrame = pd.DataFrame(), density_df: DataFrame =pd.DataFrame(), textures_df: DataFrame =pd.DataFrame(),
intervals_info: DataFrame = pd.DataFrame(),intervals_types: DataFrame =pd.DataFrame(), clefs_info: DataFrame =pd.DataFrame(), emphasised_scale_degrees_info_A: DataFrame =pd.DataFrame(),harmony_df: DataFrame = pd.DataFrame(), key_areas: DataFrame = pd.DataFrame(), chords: DataFrame = pd.DataFrame()):
    global rows_groups
    global not_used_cols
    rows_groups=rg
    not_used_cols=nuc

    visualiser_lock = True #remve with threads

    if groups:
        # if sequential:
        results_path = path.join(results_path_factorx, '_'.join(groups))
        if not os.path.exists(results_path):
            os.mkdir(results_path)
    else:
        results_path = results_path_factorx
    if os.path.exists(path.join(results_path, 'visualisations')):
        remove_folder_contents(
            path.join(results_path, 'visualisations'))
    else:
        os.makedirs(path.join(results_path, 'visualisations'))
    # MUTITHREADING
    try:
        # executor = concurrent.futures.ThreadPoolExecutor()
        # visualiser_lock = threading.Lock()
        # futures = []
        if not melody_values.empty:
            #     # futures.append(executor.submit(Melody_values, Melody_values, results_path, '-'.join(groups) + "_1Values.xlsx", sorting_lists,
            #     #                visualiser_lock, additional_info, True if i == 0 else False, groups if groups != [] else None))
            melody_values = pd.concat([common_columns_df, melody_values], axis=1)
            Melody_values(factor, _cfg, melody_values, results_path, '-'.join(groups) + str(factor) + '_factor_'+ "_MelodyValues.xlsx",
                    visualiser_lock, additional_info, True if factor == 0 else False, groups if groups != [] else None)
        if not density_df.empty:
            density_df = pd.concat(
                [common_columns_df, density_df,notes_df], axis=1)
            Densities(factor, _cfg, density_df, results_path, '-'.join(groups) + str(factor) + '_factor_'+  "_Densities.xlsx", visualiser_lock, groups if groups != [] else None, additional_info)
        if not textures_df.empty:
            textures_df = pd.concat(
            [common_columns_df, textures_df,notes_df], axis=1)
            Textures(factor, _cfg, textures_df, results_path, '-'.join(groups) + str(factor) + '_factor_'+ "_Textures.xlsx", visualiser_lock, groups if groups != [] else None, additional_info)
        if not intervals_info.empty:
            intervals_info=pd.concat([common_columns_df, intervals_info], axis=1)
            Intervals(factor, _cfg, intervals_info, '-'.join(groups) + str(factor) + '_factor_'+ "Intervals.xlsx",
                                _cfg.sorting_lists["Intervals"], results_path, visualiser_lock, additional_info, groups if groups != [] else None)
            
            absolute_intervals=make_intervals_absolute(intervals_info)
            Intervals(factor, _cfg, absolute_intervals, '-'.join(groups) + str(factor) + '_factor_'+ "Intervals_absolute.xlsx",
                             _cfg.sorting_lists["Intervals_absolute"], results_path, visualiser_lock, additional_info, groups if groups != [] else None)
        if not intervals_types.empty:
            intervals_types = pd.concat([common_columns_df, intervals_types], axis=1)
            Intervals_types(factor, _cfg, intervals_types, results_path, '-'.join(groups) + str(factor) + '_factor'+ "_Interval_types.xlsx", visualiser_lock, groups if groups != [] else None, additional_info)
        
        if not emphasised_scale_degrees_info_A.empty:
            emphasised_scale_degrees_info_A = pd.concat([common_columns_df,emphasised_scale_degrees_info_A], axis=1)
            Emphasised_scale_degrees(factor, _cfg, emphasised_scale_degrees_info_A,  _cfg.sorting_lists["ScaleDegrees"], '-'.join(
                groups) + str(factor) + '_factor'+  "_Scale_degrees.xlsx", results_path, visualiser_lock, groups if groups != [] else None, additional_info)
        # if not Emphasised_scale_degrees_info_B.empty:
        #     Emphasised_scale_degrees( Emphasised_scale_degrees_info_B, sorting_lists["ScaleDegrees"], '-'.join(
        #         groups) + "_4bScale_degrees_relative.xlsx", results_path, sorting_lists, visualiser_lock, groups if groups != [] else None, additional_info)
        if not clefs_info.empty:
            # clefs_info= pd.concat([common_columns_df,clefs_info], axis=1)
            Intervals(factor, _cfg, clefs_info, '-'.join(groups) + str(factor) + '_factor_'+  "_Clefs_in_voice.xlsx",
                             _cfg.sorting_lists["Clefs"], results_path, visualiser_lock, additional_info, groups if groups != [] else None)
        # if not harmony_df.empty:
        #     harmony_df= pd.concat([common_columns_df,harmony_df], axis=1)
        #     # Harmonic_data(cfg, harmony_df, '-'.join(groups) + "Harmonic_rythm.xlsx",
        #     #                 sorting_lists["Clefs"], results_path, sorting_lists, visualiser_lock, additional_info, groups if groups != [] else None)
        if not chords.empty:
            chords = pd.concat([common_columns_df, chords], axis=1)
            Chords(factor, _cfg, chords, results_path, '-'.join(groups) + str(factor) + '_factor_'+  "Chords.xlsx", visualiser_lock, groups if groups != [] else None, additional_info)
        # if not key_areas.empty:
        #     key_areas= pd.concat([common_columns_df,key_areas], axis=1)
        #     # clefs_info= pd.concat([common_columns_df,clefs_info], axis=1)
        #     Keyareas(cfg, key_areas, results_path, '-'.join(groups) + "_Key_areas.xlsx",
        #                 sorting_lists, visualiser_lock, groups if groups != [] else None, additional_info)
        #     Keyareas_weigthed(cfg, key_areas, results_path, '-'.join(groups) + "_Key_areas.xlsx",
        #                 sorting_lists, visualiser_lock, groups if groups != [] else None, additional_info)
            
            # wait for all
            # if sequential:
            # kwargs = {'total': len(futures), 'unit': 'it',
            #             'unit_scale': True, 'leave': True}
            # for f in tqdm(concurrent.futures.as_completed(futures), **kwargs):
            #     pass
            # else:
            #     for f in concurrent.futures.as_completed(futures):
            #         pass
    except KeyboardInterrupt:
        _cfg.write_logger.error('\033[93mAn error ocurred during the report generation process. \033[37m')
        sys.exit(2)

###########################################################################################################################################################
# Function is in charge of printing each group: 'Per Opera', 'Per City'...
#
#   sheet: the openpyxl object in which we will write
#   grouped: dataframe already grouped by the main factor and the additional information that it may show
#   row_number, column_number: row and column in which we will start to write the information
#   columns: list of the dataframe (grouped) column names that we need to access (as it doesn't necessarily correspond to the names that we want to print)
#   third_columns: list of the names of the columns that we need to print
#   computations_columns: information about the matematical computation that has to be done to each column (sum, mean...)
#   ----------------
#   first_columns: list of column names to print in first place, along with the number of columns that each has to embrace
#   second_columns: list of column names to print in second place, along with the number of columns that each has to embrace
#   per: boolean value to indicate if we need to compute the excel in absolute values or percentage (by default it is absolute)
#   average: boolean value to indicate if we want the average row at the last group's row
#   last_column: boolean value to indicate if we want a summarize on the last column
#   last_column_average: boolean to indicate if we want the last column to have each row's average (otherwise, the total is writen)
#   additional_info: number of extra columns containing additional info
#   ponderate: boolean to indicate if we want to ponderate the data printed or not
#   not_grouped_df: tuple containing the additional information columns and the dataframe without applying groupBy
#
###########################################################################################################################################################

def print_groups(sheet: ExcelWriter, grouped:DataFrame, row_number: int, column_number: int, columns: list, third_columns: list, computations_columns: list,
                 first_columns: list = None, second_columns: List=None, per: bool=False, average:bool=False, last_column: bool=False,
                 last_column_average: bool=False, additional_info: DataFrame=None, ponderate: bool=False, not_grouped_df:DataFrame=None):
    len_add_info = 0  # Space for the possible column of additional information
    if additional_info:
        len_add_info = additional_info
    # WRITE THE COLUMN NAMES (<first>, <second> and third)
    if first_columns:
        write_columns_titles_variable_length(
            sheet, row_number, column_number + 1 + len_add_info, first_columns, titles1Fill)
        row_number += 1
    if second_columns:
        write_columns_titles_variable_length(
            sheet, row_number, column_number + 1 + len_add_info, second_columns, titles2Fill)
        row_number += 1
    starting_row = row_number
    write_columns_titles(sheet, row_number, column_number +
                         1 + len_add_info, third_columns)
    row_number += 1
    exception = -1
    total_analysed_column = "Total analysed" in columns
    cnumber = column_number
    # PRINT EACH ROW
    # store each result in case of need of calculating the percentage (per = True)
    columns_values = {c: [] for c in columns}

    # Subgroup: ex: Berlin when groupping by Territory
    for s, subgroup in enumerate(grouped):

        cnumber = column_number  
        # Print row name
        if type(subgroup[0]) != tuple:  # It has no additional information
            # could be a tuple if we have grouped by more than one element
            sheet.cell(row_number, column_number).value = subgroup[0]
            sheet.cell(row_number, column_number).font =  FONT

        else:
            for g in subgroup[0]:
                sheet.cell(row_number, cnumber).value = g
                sheet.cell(row_number, cnumber).font =  FONT
                cnumber += 1

        subgroup_data = subgroup[1]
        cnumber = column_number + 1 + len_add_info

        total_analysed_row = [1]
        if not_grouped_df is not None:
            if type(subgroup[0]) != tuple:
                cond = not_grouped_df[1][not_grouped_df[0][0]] == subgroup[0]
                not_grouped_information = not_grouped_df[1][cond].drop(
                    not_grouped_df[0] + ['Total analysed'], axis=1)
            else:
                for sb in range(len(subgroup[0])):
                    cond = not_grouped_df[1][not_grouped_df[0]
                                             [sb]] == subgroup[0][sb]
                    not_grouped_information = not_grouped_df[1][cond]
                not_grouped_information = not_grouped_information.drop(
                    not_grouped_df[0] + ['Total analysed'], axis=1)
        else:
            not_grouped_information = None

        # compute each column's value for the present row (subgroup) and print it
        for i, c in enumerate(columns):
            column_computation = computations_columns[i]
            extra_info = []
            if column_computation == 'mean_density':
                extra_info = subgroup_data[c+'Measures']*subgroup_data['NumberOfBeats'] # Its a multiplication because it like a division to tjhe whole result
                value = compute_value(subgroup_data[c], column_computation, total_analysed_row,
                                      not_grouped_information, ponderate, extra_info=extra_info)
                                      
            elif column_computation == 'mean_texture':
                if not subgroup_data[c].isnull().all():
                    notes = subgroup_data[c.split('/')[0]+'Notes']
                    notes_next = subgroup_data[c.split('/')[1]+'Notes']
                    value = compute_value(notes, column_computation, total_analysed_row,
                                        not_grouped_information, ponderate, extra_info=notes_next)
                else:
                    value=0.0
            else:
                value = compute_value(subgroup_data[c], column_computation, total_analysed_row,
                                      not_grouped_information, ponderate)  # absolute value
            if c == "Total analysed":
                total_analysed_row = subgroup_data[c].to_list()
                sheet.cell(row_number, cnumber).value = value
                sheet.cell(row_number, cnumber).font =  FONT

            if c == interval.ABSOLUTE_INTERVALLIC_TRIM_RATIO:  # EXCEPTION
                sheet.cell(row_number, cnumber).value = str(
                    round(value * 100, 1)) + '%'
                sheet.cell(row_number, cnumber).font =  FONT
                cnumber += 1
                exception = i - 1
            elif not per:
                sheet.cell(row_number, cnumber).value = str(
                    value) + '%' if ponderate and column_computation == "sum" and c != "Total analysed" else str(value).replace(',', '.')
                sheet.cell(row_number, cnumber).font =  FONT
                cnumber += 1
            # store each value in case of needing to print the percentage
            columns_values[c].append(value)
        row_number += 1

    if total_analysed_column:  # We don't need Total analysed from this point
        del columns_values['Total analysed']
        computations_columns = computations_columns[1:]

    last_used_row = row_number
    if per or last_column:  # These are the two conditions in which we need to transpose valores_columnas
        # Transpose valores_columnas to valores_filas (change perspective from column to rows)
        columns_list = list(columns_values.values())
        keys_columnas = list(columns_values.keys())
        rows_values = []
        len_lists = len(columns_list[0])
        for i in range(len_lists):
            rows_values.append(round(sum([lc[i] for x, lc in enumerate(
                columns_list) if "All" not in keys_columnas[x]]), 3))

    # PRINT EACH CELL IF PER IS TRUE, now that we have all the information
    if per:
        cn = column_number + len_add_info + \
            2 if total_analysed_column else column_number + len_add_info + 1
        for i in range(len(columns_list)):  # Traverse each column's information
            row_number = starting_row
            sum_column = sum(columns_list[i]) if sum(
                columns_list[i]) != 0 else 1
            for j in range(len(columns_list[i])):
                row_number += 1
                # COMPUTE THE HORIZONTAL AVERAGE (average within the present column or row)
                if average:
                    value = round(
                        (columns_list[i][j]/rows_values[j])*100, 3)
                else:
                    value = round((columns_list[i][j]/sum_column)*100, 3)
                columns_values[keys_columnas[i]][j] = value if str(
                    value) != "nan" else 0  # update the value
                sheet.cell(row_number, cn).value = str(
                    value) + "%"  # print the value
                sheet.cell(row_number, cn).font =  FONT
            cn += 1

        cnumber = cn

    # RECALCULATE VALORES_FILAS AGAIN TO GET THE MOST UPDATED DATA
    columns_list = list(columns_values.values()
                           )  # Get the updated version
    if per:  # Compute valores_filas again
        rows_values = []
        for i in range(len_lists):
            rows_values.append(round(sum([lc[i] for x, lc in enumerate(
                columns_list) if "All" not in keys_columnas[x]]), 3))

    # PRINT THE LAST COLUMN (AVERAGE OR TOTAL)
    if last_column:
        if last_column_average:
            for j, vf in enumerate(rows_values):
                try:
                    rows_values[j]=round(vf / (len(columns_list) - len(
                [c for c in keys_columnas if "All" in c])- len([col[j] for col in columns_list if col[j] == 0.0])), 3)
                except ZeroDivisionError:
                    rows_values[j]= 0.0
     
        # PRINT THE LAST COLUMN, CONTAINING THE TOTAL OR THE AVERAGE OF THE DATA
        print_averages_total_column(sheet, starting_row, cnumber, rows_values, average=last_column_average,
                                    per=per or ponderate and all(c == "sum" for c in computations_columns))

    # PRINT LAST ROW (TOTAL OR AVERAGE)
    values=np.asarray(list(columns_values.values()))
    for i, c in enumerate(columns_values):
        # In case we have all zeros in a row means that element is not present in the aria so we don't take into account for calculations
        for row in range(0,values.shape[1]):
            if all([str(e)=='0.0' for e in values[:,row]]):
                del columns_values[c][0]

        if average:  
            columns_values[c] = compute_average(
                columns_values[c], computations_columns[i])
        else:  # total
            columns_values[c] = round(sum(columns_values[c]), 3)

    final_values = list(columns_values.values())
    # Take the last value computed for the last column (average or total)
    if last_column:
        if average:
            final_values.append(
                round(sum(rows_values) / len(rows_values), 3))
        else:
            final_values.append(round(sum(rows_values), 3))
    data_column = column_number + len_add_info + \
        2 if total_analysed_column else column_number + len_add_info + 1
    print_averages_total(sheet, last_used_row, final_values, column_number, data_column, average=average,
                         per=per or ponderate and all(c == "sum" for c in computations_columns), exception=exception)
    ###

    return last_used_row + 1, cnumber + 1

##########################################################################################################
# Function in charge of printting the data, the arguments are the same as the explained in excel_sheet  #
##########################################################################################################

def row_iteration(sheet: ExcelWriter, columns: list, row_number: int, column_number: int, data: DataFrame, third_columns: list, computations_columns: List[str], sorting_lists: list, group: list=None, first_columns: list=None, second_columns: list=None, per: bool=False, average: bool=False, last_column: bool=False, last_column_average: bool=False,
                  columns2: list=None, data2: DataFrame=None, third_columns2: list=None, computations_columns2: list=None, first_columns2: list=None, second_columns2: list=None, additional_info: list=[], ponderate: bool =False):
    all_columns = list(data.columns)
    for row in rows_groups:  # Geography, Dramma, Opera, Aria, Label, Composer...
        if row in all_columns or any(sub in all_columns for sub in rows_groups[row][0]):
            forbiden = [NAME]
            if group != None:
                forbiden = [forbiden_groups[group[i]]
                            for i in range(len(group))]
                forbiden = [item for sublist in forbiden for item in sublist]
            if group == None and row not in forbiden: #it was 'or' and not 'and' before, but change considered neccessary  

                # 1. Write the Title in Yellow
                sheet.cell(row_number, column_number).value = "Per " + row.replace('Aria', '')
                sheet.cell(row_number, column_number).font =  FONT
                sheet.cell(row_number, column_number).fill = YELLOWFILL
                row_number += 1
                sorting = rows_groups[row][1]
                # 2. Write the information depending on the subgroups (ex: Geography -> City, Country)
                if len(rows_groups[row][0]) == 0:  # No subgroups
                    starting_row = row_number
                    # Sort the dataframe based on the json sorting_lists in Json_extra
                    data = sort_dataframe(data, row, sorting_lists, sorting)
                    groups_add_info, add_info = get_groups_add_info(
                        data, row, additional_info)
                    row_number, last_column_used = print_groups(sheet, data.groupby(groups_add_info, sort=False), row_number, column_number, columns, third_columns, computations_columns, first_columns, second_columns, per=per,
                                                                average=average, last_column=last_column, last_column_average=last_column_average, additional_info=add_info, ponderate=ponderate, not_grouped_df=(groups_add_info, data[groups_add_info + columns]))
                    if columns2 != None:  # Second subgroup
                        groups_add_info, add_info = get_groups_add_info(
                            data, row, additional_info)
                        if data2 is not None:
                            data2 = sort_dataframe(
                                data2, row, sorting_lists, sorting)
                        _, _ = print_groups(sheet, data.groupby(groups_add_info, sort=False) if data2 is None else data2.groupby(groups_add_info, sort=False), starting_row, last_column_used + 1, columns2, third_columns2, computations_columns2, first_columns2,
                                            second_columns2, per=per, average=average, last_column=last_column, last_column_average=last_column_average, additional_info=add_info, ponderate=ponderate, not_grouped_df=(groups_add_info, data[groups_add_info + columns]))
                else:  # has subgroups, ex: row = Date, subgroups: Year
                    if rows_groups[row][0] == [CHARACTER, ROLE, GENDER]:
                        data_joint = data.copy()
                        data = split_voices(data)
                    for i, subrows in enumerate(rows_groups[row][0]):
                        if (subrows == None or subrows not in forbiden) and subrows in all_columns:
                            if "Tempo" in subrows:
                                data[subrows] = data[subrows].fillna('')
                            starting_row = row_number
                            sort_method = sorting[i]
                            sheet.cell(
                                row_number, column_number).value = subrows
                            sheet.cell(
                                row_number, column_number).fill = greenFill
                            sheet.cell(row_number, column_number).font =  FONT
                            data = sort_dataframe(
                                data, subrows, sorting_lists, sort_method)

                            groups_add_info, add_info = get_groups_add_info(
                                data, subrows, additional_info)
                            row_number, last_column_used = print_groups(sheet, data.groupby(groups_add_info, sort=False), row_number, column_number, columns, third_columns, computations_columns, first_columns, second_columns, per=per,
                                                                        average=average, last_column=last_column, last_column_average=last_column_average, additional_info=add_info, ponderate=ponderate, not_grouped_df=(groups_add_info, data[groups_add_info + columns]))
                            if columns2 != None:  # Second subgroup
                                if "Tempo" in subrows and data2 is not None:
                                    data2[subrows] = data2[subrows].fillna('')
                                if data2 is not None:
                                    data2 = sort_dataframe(
                                        data2, subrows, sorting_lists, sort_method)
                                groups_add_info, add_info = get_groups_add_info(
                                    data, subrows, additional_info)
                                _, _ = print_groups(sheet, data.groupby(groups_add_info, sort=False) if data2 is None else data2.groupby(groups_add_info, sort=False), starting_row, last_column_used + 1, columns2, third_columns2, computations_columns2, first_columns2,
                                                    second_columns2, per=per, average=average, last_column=last_column, last_column_average=last_column_average, additional_info=add_info, ponderate=ponderate, not_grouped_df=(groups_add_info, data[groups_add_info + columns]))

                            row_number += 1
                    if rows_groups[row][0] == [CHARACTER, ROLE, GENDER]:
                        data = copy.deepcopy(data_joint)
                row_number += 1
    return row_number


def excel_sheet(sheet: ExcelWriter, columns: list, data: DataFrame, third_columns: list, computations_columns: list, sorting_lists: list, groups: list=None, first_columns: list=None, second_columns: list=None, per: bool=False, average: bool=False, last_column: bool=False, last_column_average: bool=False,
                 columns2: list=None, data2: DataFrame=None, third_columns2: list=None, computations_columns2: list=None, first_columns2: list=None, second_columns2: list=None, additional_info: list=[], ponderate: bool=False):
                # second_subgroup = False, second_subgroup_info = {}, third_subgroup = False, third_subgroup_info = {}, fourth_subgroup = False, fourth_subgroup_info = {},fifth_subgroup=False, fifth_subgroup_info={}, six_subgroup=False, six_subgroup_info={},
                # additional_info = [], ponderate = False, total = True, valores_filas = [], want_total_counts = False):
    
    row_number = 1  # we start writing in row 1
    column_number = 1
    if groups == None:
        row_iteration(sheet, columns, row_number, column_number, data, third_columns, computations_columns, sorting_lists, first_columns=first_columns, second_columns=second_columns, per=per,
                      average=average, last_column=last_column, last_column_average=last_column_average, columns2=columns2, data2=data2, third_columns2=third_columns2, computations_columns2=computations_columns2, first_columns2=first_columns2, second_columns2=second_columns2, additional_info=additional_info, ponderate=ponderate)
    else:
        # we may be grouping by more than 2 factors
        data_grouped = data.groupby(list(groups))

        last_printed = {i: ('', 0) for i in range(len(groups))}
        for group, group_data in data_grouped:
            cnumber = column_number
            group = [group] if type(group) != tuple else group
            for i, g in enumerate(group):
                if last_printed[i][0] != g:
                    sheet.cell(row_number, cnumber).value = g
                    sheet.cell(row_number, cnumber).fill = factors_Fill[i]
                    sheet.cell(row_number, column_number).font =  BOLD
                    counter_g = data[groups[i]].tolist().count(g)
                    sheet.cell(row_number, cnumber + 1).value = counter_g
                    if last_printed[i][0] != '':
                        sheet.merge_cells(
                            start_row=last_printed[i][1], start_column=i + 1, end_row=row_number - 2, end_column=i + 1)
                        sheet.cell(last_printed[i][1],
                                  i + 1).fill = factors_Fill[i]

                    last_printed[i] = (g, row_number + 1)

                row_number += 1
                cnumber += 1
            data2_grouped = None
            if data2 is not None:
                data2_grouped = data2.groupby(list(groups)).get_group(
                    group if len(group) > 1 else group[0])
            rn = row_iteration(sheet, columns, row_number, cnumber, group_data, third_columns, computations_columns, sorting_lists, group=groups, first_columns=first_columns, second_columns=second_columns, per=per,
                               average=average, last_column=last_column, last_column_average=last_column_average, columns2=columns2, data2=data2_grouped, third_columns2=third_columns2, computations_columns2=computations_columns2, first_columns2=first_columns2, second_columns2=second_columns2, additional_info=additional_info, ponderate=ponderate)
            row_number = rn
        # merge last cells
        for i, g in enumerate(group):
            if last_printed[i][0] == g:
                sheet.merge_cells(
                    start_row=last_printed[i][1], start_column=i + 1, end_row=row_number - 2, end_column=i + 1)
                sheet.cell(last_printed[i][1],  i + 1).fill = factors_Fill[i]

### MELODY ###

########################################################################
# Function ment to write the Melody_values excel
# -------
# data: Melody_values dataframe
# results_path: where to store the data
# name: name that the excel will take
# sorting lists: lists that will be used for sorting the results
# visualiser_lock: lock used to avoid deadlocks, as matplotlib is not thread safe
# additional info: columns additional to each
# remove_columns: used for factor 0, to avoid showing unusefull information
# groups: used for factor > 1
########################################################################

def Melody_values(factor, _cfg: Configuration, data: DataFrame, results_path: str, name: str, visualiser_lock: Lock, additional_info: list=[], remove_columns: bool=False, groups: list=None):
    try:
        workbook = openpyxl.Workbook()
        data.rename(columns={interval.INTERVALLIC_MEAN: "Intervallic ratio", interval.TRIMMED_ABSOLUTE_INTERVALLIC_MEAN: "Trimmed intervallic ratio", interval.ABSOLUTE_INTERVALLIC_TRIM_DIFF: "dif. Trimmed",
                             interval.ABSOLUTE_INTERVALLIC_MEAN: "Absolute intervallic ratio", interval.INTERVALLIC_STD: "Std", interval.ABSOLUTE_INTERVALLIC_STD: "Absolute Std", interval.ABSOLUTE_INTERVALLIC_TRIM_RATIO: "% Trimmed"}, inplace=True)
        data.rename(columns={ambitus.LOWEST_INDEX: "LowestIndex", ambitus.HIGHEST_INDEX: "HighestIndex", ambitus.HIGHEST_MEAN_INDEX: "HighestMeanIndex", ambitus.LOWEST_MEAN_INDEX: "LowestMeanIndex",
             ambitus.LOWEST_NOTE: "LowestNote", ambitus.LOWEST_MEAN_NOTE: "LowestMeanNote",ambitus.HIGHEST_MEAN_NOTE: "HighestMeanNote", ambitus.LOWEST_MEAN_INDEX: "LowestMeanIndex",ambitus.HIGHEST_NOTE: "HighestNote" }, inplace=True)
        # HOJA 1: STATISTICAL_VALUES
        column_names = ["Total analysed", "Intervallic ratio", "Trimmed intervallic ratio", "dif. Trimmed",
                        "% Trimmed", "Absolute intervallic ratio", "Std", "Absolute Std"]

        if lyrics.SYLLABIC_RATIO in data.columns:
            data.rename(columns={lyrics.SYLLABIC_RATIO: 'Syllabic ratio'}, inplace=True)
            column_names.append('Syllabic ratio')

        # HAREMOS LA MEDIA DE TODAS LAS COLUMNAS
        computations = ['sum'] + ["mean"]*(len(column_names) - 1)
        excel_sheet(workbook.create_sheet("Statistical_values"), column_names, data, column_names, computations,
                    _cfg.sorting_lists, groups=groups, average=True, additional_info=additional_info, ponderate=True)

        # HOJA 2: AMBITUS
        first_column_names = [("", 1), ("Lowest", 2), ("Highest", 2), ("Lowest", 2), ("Highest", 2), (
            "Ambitus", 6)] if not remove_columns else [("", 1), ("Lowest", 2), ("Highest", 2), ("Ambitus", 2)]

        second_column_names = [("", 5), ("Mean", 2), ("Mean", 2), ("Largest", 2), ("Smallest", 2), ("Mean", 2)] if not remove_columns else [("", 5), ("Largest", 2)]

        third_columns_names = ["Total analysed", "Note", "Index", "Note", "Index", "Note", "Index", "Note", "Index", "Semitones", "Interval", "Semitones", "Interval",
         "Semitones", "Interval"] if not remove_columns else ["Total analysed", "Note", "Index", "Note", "Index", "Semitones", "Interval"]

        computations = ["sum", "minNote", "min", "maxNote", "max", 'meanNote', 'mean', 'meanNote', 'mean', 'max', "maxInterval", 'min', "minInterval", 'absolute',
                        'absoluteInterval', "meanSemitones", "meanInterval"] if not remove_columns else ["sum", "minNote", "min", "maxNote", "max", 'max', "maxInterval"]

        columns = columns_alike_our_data(
            third_columns_names, second_column_names, first_column_names)

        excel_sheet(workbook.create_sheet("Ambitus"), columns, data, third_columns_names, computations, _cfg.sorting_lists, groups=groups,
                     first_columns=first_column_names, second_columns=second_column_names, average=True, additional_info=additional_info)

        # HOJA 3: LARGEST_LEAPS
        second_column_names = [("", 1), ("Ascending", 2), ("Descending", 2)]
        third_columns_names = ["Total analysed",
                               "Semitones", "Interval", "Semitones", "Interval"]
        columns = columns_alike_our_data(
            third_columns_names, second_column_names)

        computations = ["sum", "max", "maxInterval", "min", "minInterval"]

        data.rename(columns={interval.LARGEST_ASC_INTERVAL: "AscendingInterval",interval.LARGEST_ASC_INTERVAL_SEMITONES: "AscendingSemitones", interval.LARGEST_DESC_INTERVAL: "DescendingInterval",interval.LARGEST_DESC_INTERVAL_SEMITONES: "DescendingSemitones"}, inplace=True)

        excel_sheet(workbook.create_sheet("Largest_leaps"), columns, data, third_columns_names, computations,
                     _cfg.sorting_lists, groups=groups, second_columns=second_column_names, average=True, additional_info=additional_info)

        if "Sheet" in workbook.get_sheet_names():  # Delete the default sheet
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)

        adjust_excel_width_height(workbook)


        workbook.save(os.path.join(results_path, name))

        # with visualiser_lock:
        # VISUALISATIONS
        columns_visualisation = [
            'Intervallic ratio', 'Trimmed intervallic ratio',  'Std', "Absolute intervallic ratio","Absolute Std"]
        if groups:
            data_grouped = data.groupby(list(groups))
            for g, datag in data_grouped:
                result_visualisations = path.join(
                    results_path, 'visualisations', g)
                if not os.path.exists(result_visualisations):
                    os.mkdir(result_visualisations)

                name_bar = path.join(
                    result_visualisations, name.replace('.xlsx', '.png'))
                melody_bar_plot(
                    name_bar, datag, columns_visualisation, second_title=str(g))
                name_box = path.join(
                    result_visualisations, 'Ambitus' + name.replace('.xlsx', '.png'))
                box_plot(name_box, datag, second_title=str(g))

        elif factor == 1:
            groups = [i for i in rows_groups]
            for row in rows_groups:
                plot_name = name.replace(
                    '.xlsx', '') + 'Per ' + str(row.replace('Aria','').upper()) + '.png'
                name_bar =path.join(results_path,'visualisations', plot_name)
                if row not in not_used_cols:
                    if len(rows_groups[row][0]) == 0:  # no sub-groups
                        data_grouped = data.groupby(row, sort=True)
                        if data_grouped:
                            melody_bar_plot(name_bar, data_grouped, columns_visualisation, second_title='Per ' + str(row.replace('Aria','').upper()))
                            name_box = path.join(results_path,'visualisations', 'Ambitus'+'Per_' + str(row.upper())+name.replace('.xlsx', '.png'))
                            if row == CLEF1: #Exception for boxplots
                                box_plot(name_box, data_grouped, second_title='Per '+ str(row.replace('Aria','').upper()))
                    else: #subgroups
                            for i, subrow in enumerate(rows_groups[row][0]):
                                if subrow not in EXCEPTIONS:
                                    plot_name = name.replace(
                                        '.xlsx', '') + '_Per_' + str(row.upper()) + '_' + str(subrow) + '.png'
                                    name_bar = results_path + '\\visualisations\\' + plot_name
                                    data_grouped = data.groupby(subrow)
                                    melody_bar_plot(name_bar, data_grouped, columns_visualisation, second_title='Per ' + str(subrow.replace('Aria','').upper()))
                                    name_box = path.join(
                                    results_path, 'visualisations', 'Ambitus' + name.replace('.xlsx', '.png'))
                                    
                                    if subrow == ROLE:
                                        box_plot(name_box, data_grouped, second_title='Per '+ str(subrow.replace('Aria','').upper()))
        else:                   
            name_bar = path.join(results_path,path.join('visualisations', name.replace('.xlsx', '.png')))
            melody_bar_plot(name_bar, data, columns_visualisation)
            name_box = path.join(
                results_path, 'visualisations', 'Ambitus' + name.replace('.xlsx', '.png'))
            box_plot(name_box, data)
    except Exception as e:
        _cfg.write_logger.info('{}  Problem found: {}'.format(name, e))


def Intervals(factor, _cfg: Configuration, data: DataFrame, name: str, sorting_list: list, results_path: str, visualiser_lock: Lock, additional_info: list=[], groups: list=None):
    try:
        workbook = openpyxl.Workbook()
        all_columns = data.columns.tolist()
        general_cols = copy.deepcopy(not_used_cols)
        for row in rows_groups:
            if len(rows_groups[row][0]) == 0:
                general_cols.append(row)
            else:
                general_cols += rows_groups[row][0]

        # nombres de todos los intervalos
        third_columns_names_origin = set(all_columns) - set(general_cols)
        third_columns_names_origin = sort(
            third_columns_names_origin, sorting_list)
        third_columns_names = ['Total analysed'] + third_columns_names_origin

        # esta sheet va de sumar, así que en todas las columnas el cómputo que hay que hacer es sumar!
        computations = ["sum"]*len(third_columns_names)

        excel_sheet(workbook.create_sheet("Weighted"), third_columns_names, data, third_columns_names, computations, _cfg.sorting_lists,
                     groups=groups, average=True, last_column=True, last_column_average=False, additional_info=additional_info, ponderate=True)
        if factor>=1:
            excel_sheet(workbook.create_sheet("Horizontal Per"), third_columns_names, data, third_columns_names, computations, _cfg.sorting_lists,
                     groups=groups, per=True, average=True, last_column=True, last_column_average=False, additional_info=additional_info)


        if "Sheet" in workbook.get_sheet_names():  # Delete the default sheet
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)
        adjust_excel_width_height(workbook)
            
        workbook.save(os.path.join(results_path, name))

        # with visualiser_lock:
        # VISUALISATIONS
        if 'Clefs' in name:
            title = 'Use of clefs in the vocal part'
        elif 'Intervals_absolute' in name:
            title = 'Presence of intervals (direction dismissed)'
        else:
            title = 'Presence of intervals (ascending and descending)'

        if groups:
            data_grouped = data.groupby(list(groups))
            for g, datag in data_grouped:
                result_visualisations = path.join(
                    results_path, 'visualisations', g)
                if not os.path.exists(result_visualisations):
                    os.mkdir(result_visualisations)
                name_bar = path.join(
                    result_visualisations, name.replace('.xlsx', '.png'))
                bar_plot(name_bar, datag, third_columns_names_origin,
                            'Intervals' if 'Clef' not in name else 'Clefs', title, second_title=str(g))
        elif factor == 1:
            groups = [i for i in rows_groups]
            for row in rows_groups:
                if row not in not_used_cols:
                    name_bar = path.join(results_path, 'visualisations',
                                    name.replace('.xlsx',  '').replace('{0}_factor_'.format(str(factor)),'') + '_Per_' + str(row.replace('Aria','').upper()) + '.png')
                    if len(rows_groups[row][0]) == 0:  # no sub-groups
                        data_grouped = data.groupby(row, sort=True)
                        if data_grouped:
                            bar_plot(name_bar, data_grouped, third_columns_names_origin,
                                        'Intervals' + '\n' + str(row).replace('Aria','').upper() if 'Clef' not in name else 'Clefs' + str(row).replace('Aria','').upper(), title)
                    else: #subgroups
                        for i, subrow in enumerate(rows_groups[row][0]):
                            if subrow not in EXCEPTIONS:
                                name_bar = path.join(results_path, 'visualisations',
                                    name.replace('.xlsx',  '').replace('{0}_factor_'.format(str(factor)),'') + '_Per_' + str(subrow.replace('Aria','').upper()) + '.png')
                                data_grouped = data.groupby(subrow)
                                bar_plot(name_bar, data_grouped, third_columns_names_origin,
                                        'Intervals' + str(row).replace('Aria','').upper() if 'Clef' not in name else 'Clefs' + str(row).replace('Aria','').upper(), title)
                                # melody_bar_plot(name_bar, data_grouped, columns_visualisation, second_title='Per ' + str(subrow.replace('Aria','').upper()))
                                name_box = path.join(
                                results_path, 'visualisations', 'Ambitus' + name.replace('.xlsx', '.png'))

        else:
            name_bar = path.join(
                results_path, 'visualisations', name.replace('.xlsx', '.png'))
            bar_plot(name_bar, data, third_columns_names_origin,
                        'Intervals' if 'Clef' not in name else 'Clefs', title)
    except Exception as e:
        _cfg.write_logger.info('{}  Problem found: {}'.format(name, e))

#########################################################
# Function to generate the reports file Intervals_types.xlsx  #
#########################################################

def Intervals_types(factor, _cfg: Configuration, data: DataFrame, results_path: str, name: str, visualiser_lock: Lock, groups=None, additional_info: list=[]):
    try:
        data.columns=[c.replace('Desc', 'Descending').replace('Asc', 'Ascending') for c in data.columns]
        workbook = openpyxl.Workbook()
        second_column_names = [("", 2), ("Leaps", 3), ("StepwiseMotion", 3)]
        second_column_names2 = [('', 1), ("Perfect", 3), ("Major", 3),
                                ("Minor", 3), ("Augmented", 3), ("Diminished", 3)]
        third_columns_names = ['Total analysed', "RepeatedNotes",
                               "Ascending", "Descending", "All", "Ascending", "Descending", "All"]
        third_columns_names2 = ['Total analysed', "Ascending", "Descending", "All", "Ascending", "Descending", "All",
                                "Ascending", "Descending", "All", "Ascending", "Descending", "All", "Ascending", "Descending", "All"]
        computations = ["sum"]*len(third_columns_names)
        computations2 = ['sum']*len(third_columns_names2)
        columns = columns_alike_our_data(
            third_columns_names, second_column_names)
        columns2 = columns_alike_our_data(
            third_columns_names2, second_column_names2)

        excel_sheet(workbook.create_sheet("Weighted"),
         columns, data, third_columns_names,
          computations, _cfg.sorting_lists, groups=groups, 
          last_column=True, last_column_average=False, second_columns=second_column_names,
           average=True,
                     columns2=columns2,  third_columns2=third_columns_names2, computations_columns2=computations2, second_columns2=second_column_names2, additional_info=additional_info, ponderate=True)
        if factor>=1:
            excel_sheet(workbook.create_sheet("Horizontal Per"), columns, data, third_columns_names, computations, _cfg.sorting_lists, groups=groups, second_columns=second_column_names, per=True, average=True, last_column=True, last_column_average=False,
                     columns2=columns2,  third_columns2=third_columns_names2, computations_columns2=computations2, second_columns2=second_column_names2, additional_info=additional_info)

        # borramos la sheet por defecto
        if "Sheet" in workbook.get_sheet_names():
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)
        adjust_excel_width_height(workbook)
        workbook.save(os.path.join(results_path, name))

        # with visualiser_lock:
        # VISUALISATIONS
        if groups:
            data_grouped = data.groupby(list(groups))
            for g, datag in data_grouped:
                result_visualisations = path.join(
                    results_path, 'visualisations', g)
                if not os.path.exists(result_visualisations):
                    os.mkdir(result_visualisations)

                name_cakes = path.join(result_visualisations,
                                    name.replace('.xlsx',  '') + '_AD.png')
                pie_plot(name_cakes, datag, second_title=str(g))
                name_bars = path.join(result_visualisations,
                                    name.replace('.xlsx',  '.png'))
                double_bar_plot(name_bars, data, second_title=str(g))

        elif factor == 1:
            groups = [i for i in rows_groups]
            for row in rows_groups:
                name_cakes = path.join(results_path, 'visualisations',
                                name.replace('.xlsx', '').replace('1_factor','') + '_Per_' + str(row.replace('Aria','').upper()) + '_AD.png')
                name_bars = path.join(results_path, 'visualisations',
                                name.replace('.xlsx',  '').replace('1_factor','') + '_Per_' + str(row.replace('Aria','').upper()) + '.png')
                if row not in not_used_cols:
                    if len(rows_groups[row][0]) == 0:  # no sub-groups
                        data_grouped = data.groupby(row, sort=True)
                        if data_grouped:
                            # melody_bar_plot(name_bar, data_grouped, columns_visualisation, second_title='Per ' + str(subrow.replace('Aria','').upper()))

                            pie_plot(name_cakes, data_grouped, second_title='Per ' + str(row.replace('Aria','').upper()))
                            double_bar_plot(name_bars, data_grouped, 'Per ' + str(row.replace('Aria','').upper()))

                    else: #subgroups
                            for i, subrow in enumerate(rows_groups[row][0]):
                                if subrow not in EXCEPTIONS:
                                    name_cakes = path.join(results_path, 'visualisations',
                                    name.replace('.xlsx', '') + '_Per_' + str(row.upper()) + '_' + str(subrow) + '_AD.png')
                                    name_bars = path.join(results_path, 'visualisations',
                                    name.replace('.xlsx',  '') + '_Per_' + str(row.upper()) + '_' + str(subrow) + '.png')

                                    data_grouped = data.groupby(subrow)
                                    pie_plot(name_cakes, data_grouped, second_title='Per ' + str(row.replace('Aria','').upper()))
                                    double_bar_plot(name_bars, data_grouped, 'Per ' + str(row.replace('Aria','').upper()))

        else:
            name_cakes = path.join(results_path, 'visualisations',
                                name.replace('.xlsx', '') + '_AD.png')
            pie_plot(name_cakes, data)
            name_bars = path.join(results_path, 'visualisations',
                                name.replace('.xlsx',  '.png'))
            double_bar_plot(name_bars, data)
    except Exception as e:
        _cfg.write_logger.error('{}  Problem found: {}'.format(name, e))

########################################################################
# Function to generate the reports files Emphasised_scale_degrees.xlsx
########################################################################

def Emphasised_scale_degrees(factor, _cfg: Configuration, data: DataFrame, sorting_list: list, name: str, results_path: str, visualiser_lock: Lock, groups: list=None, additional_info=[]):
    try:
        workbook = openpyxl.Workbook()
        all_columns = data.columns.tolist()
        general_cols = copy.deepcopy(not_used_cols)
        for row in rows_groups:
            if len(rows_groups[row][0]) == 0:
                general_cols.append(row)
            else:
                general_cols += rows_groups[row][0]

        # nombres de todos los intervalos
        third_columns_names_origin = list(set(all_columns) - set(general_cols))
        third_columns_names_origin = sort(
            third_columns_names_origin, sorting_list)
        third_columns_names = ['Total analysed'] + third_columns_names_origin
        third_columns_names2 = ['Total analysed'] + \
            ['1', '4', '5', '7', 'Others']

        # esta sheet va de sumar, así que en todas las columnas el cómputo que hay que hacer es sumar!
        computations = ["sum"]*len(third_columns_names)
        computations2 = ["sum"]*len(third_columns_names2)

        emph_degrees = pd.DataFrame(prepare_data_emphasised_scale_degrees_second(
            data, third_columns_names, third_columns_names2))
        data2 = pd.concat(
            [data[[gc for gc in general_cols if gc in all_columns]], emph_degrees], axis=1)
        _, unique_columns = np.unique(data2.columns, return_index=True)
        data2 = data2.iloc[:, unique_columns]

        excel_sheet(workbook.create_sheet("Weighted"), third_columns_names, data, third_columns_names, computations, _cfg.sorting_lists, groups=groups, last_column=True, last_column_average=False, average=True,
                     columns2=third_columns_names2,  data2=data2, third_columns2=third_columns_names2, computations_columns2=computations2, additional_info=additional_info, ponderate=True)
        if factor>=1:
            excel_sheet(workbook.create_sheet("Horizontal Per"), third_columns_names, data, third_columns_names, computations, _cfg.sorting_lists, groups=groups, per=True, average=True, last_column=True, last_column_average=False,
                     columns2=third_columns_names2,  data2=data2, third_columns2=third_columns_names2, computations_columns2=computations2, additional_info=additional_info)

        # Delete the default sheet
        if "Sheet" in workbook.get_sheet_names():
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)
        adjust_excel_width_height(workbook)
        workbook.save(os.path.join(results_path, name))

        # with visualiser_lock:
        subtitile = 'in relation to the global key' if '4a' in name else 'in relation to the local key'
            # VISUALISATIONS
        if groups:
            data_grouped = data.groupby(list(groups))
            for g, datag in data_grouped:
                result_visualisations = path.join(
                    results_path, 'visualisations', g)
                if not os.path.exists(result_visualisations):
                    os.mkdir(result_visualisations)

                name1 = path.join(
                    result_visualisations, 'Scale_degrees_GlobalKey.png' if 'A' in name else 'Scale_degrees_LocalKey.png')
                customized_plot(
                    name1, data, third_columns_names_origin, subtitile, second_title=g)
        else:
            name1 = path.join(results_path, 'visualisations',
                                'Scale_degrees_GlobalKey.png' if 'A' in name else 'Scale_degrees_LocalKey.png')
            customized_plot(
                name1, data, third_columns_names_origin, subtitile)

    except Exception as e:
        _cfg.write_logger.info('{}  Problem found: {}'.format(name, e))


### COMMON INSTRUMENTS tasks ###

### Function that prints densities report task ###

def Densities(factor, _cfg: Configuration, data: DataFrame, results_path: str, name: str, visualiser_lock: Lock, groups: list=None, additional_info: list=[]):
    try:
        workbook = openpyxl.Workbook()

        # Splitting the dataframes to reorder them
        data_general = data[metadata_columns + ['Total analysed']].copy()
        data = data[set(data.columns).difference(metadata_columns)]
        data_general = data_general.dropna(how='all', axis=1)
        data = data.dropna(how='all', axis=1)
        density_df =data[[i for i in data.columns if i.endswith('SoundingDensity')]]
        notes_and_measures =data[[i for i in data.columns if i.endswith(
            'Measures') or i.endswith('Notes') or i.endswith('Mean') or i == 'NumberOfBeats']]

        density_df.columns = [i.replace('_','').replace('Part', '').replace(
            'Sounding', '').replace('Density', '').replace('Family', '').replace('Sound', '').replace('Notes', '') for i in density_df.columns]
        notes_and_measures.columns = [i.replace('_','').replace('Part', '').replace('SoundingMeasures', 'Measures').replace(
            'Sound', '').replace('Notes', '').replace('Mean', '').replace('Family', '') for i in notes_and_measures.columns]

        cols = sort(density_df.columns.tolist(), [
                    i.capitalize() for i in _cfg.sorting_lists['InstrumentSorting']])
        
        density_df = density_df[cols]
        third_columns_names = density_df.columns.to_list()

        second_column_names = [("", 1), ("Density", len(third_columns_names))]
        third_columns_names.insert(0, 'Total analysed')
        data = pd.concat([data_general, density_df], axis=1)
        data_total = pd.concat([data_general, notes_and_measures], axis=1)

        computations = ["sum"] + ["mean"] * (len(third_columns_names)-1)
        computations2 = ["sum"] + ["mean_density"] * \
            (len(third_columns_names)-1)
        columns = third_columns_names
        excel_sheet(workbook.create_sheet("Weighted"), columns, data, third_columns_names, computations, _cfg.sorting_lists, groups=groups, last_column=True,
                     last_column_average=True, second_columns=second_column_names, average=True, additional_info=additional_info, ponderate=False)
        if factor >=1:
            excel_sheet(workbook.create_sheet("Horizontal"), columns, data_total, third_columns_names, computations2,  _cfg.sorting_lists, groups=groups,
                     second_columns=second_column_names, per=False, average=True, last_column=True, last_column_average=True, additional_info=additional_info)

        # Deleting default sheet
        if "Sheet" in workbook.get_sheet_names():
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)
        adjust_excel_width_height(workbook)
        workbook.save(os.path.join(results_path, name))

        # with visualiser_lock: #Apply when threads are usedwith visualizer_lock=threading.Lock()
        columns.remove('Total analysed')
        title = 'Instrumental densities'
        # VISUALISATIONS
        if groups:
            data_grouped = data.groupby(list(groups))
            for g, datag in data_grouped:
                result_visualisations = results_path + \
                    '\\visualisations\\' + str(g.replace('/', '_'))
                if not os.path.exists(result_visualisations):
                    os.mkdir(result_visualisations)
                name_bar = result_visualisations + \
                    '\\' + name.replace('.xlsx', '.png')
                bar_plot_extended(name_bar, datag, columns, 'Density',
                                  'Density', title, second_title=str(g))

        elif factor == 1:  # 
            groups = [i for i in rows_groups]
            for row in rows_groups:
                plot_name = name.replace(
                    '.xlsx', '') + '_Per_' + str(row.upper()) + '.png'
                name_bar = results_path + '\\visualisations\\' + plot_name
                if row not in not_used_cols:
                    if len(rows_groups[row][0]) == 0:  # no sub-groups
                        data_grouped = data.groupby(row, sort=True)
                        if data_grouped:
                            line_plot_extended(
                            name_bar, data_grouped, columns, 'Instrument', 'Density', title, second_title='Per ' + str(row))
                    else:
                        for i, subrow in enumerate(rows_groups[row][0]):
                            if subrow not in EXCEPTIONS:
                                plot_name = name.replace(
                                    '.xlsx', '') + '_Per_' + str(row.upper()) + '_' + str(subrow) + '.png'
                                name_bar = results_path + '\\visualisations\\' + plot_name
                                data_grouped = data.groupby(subrow)
                                line_plot_extended(
                                    name_bar, data_grouped, columns, 'Instrument', 'Density', title, second_title= 'Per ' + str(subrow))
        else:
            for instr in third_columns_names:
                columns=data['AriaName']
                name_bar = results_path + '\\visualisations\\' + instr + \
                    name.replace('.xlsx', '.png')
                bar_plot_extended(name_bar, data, columns,
                                'Density', 'Density', title + ' ' + instr, instr=instr)
    except Exception as e:
        _cfg.write_logger.error('{}  Problem found: {}'.format(name, e))

### Function that prints textures report task ###

def Textures(factor, _cfg: Configuration, data: DataFrame, results_path: str, name: str, visualiser_lock: Lock, groups: list=None, additional_info: list=[]):
    try:
        workbook = openpyxl.Workbook()
        # Splitting the dataframes to reorder them
        data_general = data[metadata_columns + ['Total analysed']].copy()
        data_general = data_general.dropna(how='all', axis=1)
        notes_df = data[[i for i in data.columns if i.endswith('Notes') or i.endswith('Mean')]]
        textures_df = data[[c for c in data.columns if c.endswith('Texture')]]
        third_columns_names = []
        # Pre-treating data columns
        notes_df.columns = [i.replace('_', '').replace('Sound','').replace('Family','').replace('Part','').replace('Mean', '') for i in notes_df.columns]
        textures_df.columns = [i.replace('__Texture','').replace('__', '/').replace('Part', '') for i in textures_df.columns]
        
        for column in textures_df.columns:
            if column not in _cfg.sorting_lists["TexturesSorting"]:
                textures_df.drop([column], axis=1, inplace=True)
            else:
                notes_df[column]=textures_df[column]

        third_columns_names = _cfg.sorting_lists['TexturesSorting']

        for col in third_columns_names:
            if col not in textures_df.columns:
                textures_df[col]=np.nan
                notes_df[col]=np.nan
                notes_df[col.split('/')[0]+'_Notes']=np.nan

        cols = sort(textures_df.columns.tolist(), _cfg.sorting_lists['TexturesSorting'])  

        textures_df = textures_df[cols]
        
        # second_column_names = [("", 1), ("Textures", len(third_columns_names))]
        third_columns_names.insert(0, 'Total analysed')
        second_column_names = [
            ("", 1), ("Texture", len(third_columns_names))]

        computations = ["sum"] + ["mean"] * (len(third_columns_names)-1)
        computations2 = ["sum"] + ["mean_texture"] * \
            (len(third_columns_names)-1)
        columns = third_columns_names

        data = pd.concat([data_general, textures_df], axis=1)
        notes_df = pd.concat([data_general, notes_df], axis=1)

        excel_sheet(workbook.create_sheet("Weighted_textures"), columns, data, third_columns_names, computations, _cfg.sorting_lists, groups=groups,
                     last_column=True, last_column_average=True, second_columns=second_column_names, average=True, additional_info=additional_info, ponderate=False)
        if factor >=1:
            excel_sheet(workbook.create_sheet("Horizontal_textures"), columns, notes_df, third_columns_names, computations2,   _cfg.sorting_lists, groups=groups,
                     second_columns=second_column_names, per=False, average=True, last_column=True, last_column_average=True, additional_info=additional_info)

        # borramos la sheet por defecto
        if "Sheet" in workbook.get_sheet_names():
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)
            
        adjust_excel_width_height(workbook)
        workbook.save(os.path.join(results_path, name))

        # with visualiser_lock:
        title = 'Textures'
        columns.remove('Total analysed')
        # VISUALISATIONS
        if groups:
            data_grouped = data.groupby(list(groups))
            for g, datag in data_grouped:
                result_visualisations = results_path + \
                    '\\visualisations\\' + str(g)
                if not os.path.exists(result_visualisations):
                    os.mkdir(result_visualisations)
                name_bar = result_visualisations + \
                    '\\' + name.replace('.xlsx', '.png')
                bar_plot_extended(
                    name_bar, datag, columns, 'Instrumental Textures', title, second_title=str(g))

        elif factor == 1:
            groups = [i for i in rows_groups]
            for row in rows_groups:
                plot_name = name.replace(
                    '.xlsx', '') + '_Per_' + str(row.upper()) + '.png'
                name_bar = results_path + '\\visualisations\\' + plot_name
                if row not in not_used_cols:
                    if len(rows_groups[row][0]) == 0:  # no sub-groups
                        data_grouped = data.groupby(row, sort=True)
                        line_plot_extended(
                            name_bar, data_grouped, columns, 'Texture', 'Ratio', title, second_title='Per ' + str(row))
                    else:
                        for i, subrow in enumerate(rows_groups[row][0]):
                            if subrow not in EXCEPTIONS:
                                plot_name = name.replace(
                                    '.xlsx', '') + '_Per_' + str(row.upper()) + '_' + str(subrow) + '.png'
                                name_bar = results_path + '\\visualisations\\' + plot_name
                                data_grouped = data.groupby(subrow)
                                line_plot_extended(
                                    name_bar, data_grouped, columns, 'Texture', 'Ratio', title, second_title='Per ' + str(subrow))
        else:
            for instr in third_columns_names:
                if not all(data[instr].isnull()):
                    columns=data['AriaName']
                    name_bar = results_path + '\\visualisations\\' + instr.replace('/', '_') + \
                        name.replace('.xlsx', '.png')
                    bar_plot_extended(name_bar, data, columns,
                                        'Instrumental Textures', 'Ratio', title + ' ' +instr, instr=instr)
    except Exception as e:
            _cfg.write_logger.info('{}  Problem found: {}'.format(name, e))

### HARMONY ###

### Function that prints harmonic data report task ###

# def Harmonic_data(data, total_sections):
def Harmonic_data(_cfg: Configuration, data: DataFrame, results_path: str, name: str, sorting_lists: list, visualiser_lock: Lock, groups: list=None, additional_info: list=[]):
    # name = "Harmonic_data.xlsx"

    workbook = openpyxl.Workbook()

    num_chordTypes = sort([c.replace('Chord types', '') for c in data.columns if 'Chord types' in c], sorting_lists['ChordTypeGrouppingSorting'])
    num_additions = [c.replace('Additions', '') for c in data.columns if 'Additions' in c]
    second_column_names = [('Harmonic rhythm', 3 + len(total_sections)), ('Modulations', len(num_modulations)), ('Numerals', len(num_numerals)), ('Chord types', len(num_chordTypes)), ('Additions', len(num_additions))]
    third_column_names = ['Average', 'Voice', 'No_voice'] + total_sections + num_modulations + num_numerals + num_chordTypes + num_additions
    column_names = columns_alike_our_data(third_column_names, second_column_names)
    second_column_names = [('', 1)] + second_column_names
    third_column_names = ["Total analysed"] +third_column_names
    column_names = ["Total analysed"] + column_names
    computations = ['sum']*len(column_names)

    excel_sheet(workbook.create_sheet("Counts"), column_names, data, third_column_names, computations, second_columns=second_column_names,average=False, additional_info = additional_info, total = False, want_total_counts = False)
    
    #borramos la hoja por defecto
    if "Sheet" in workbook.get_sheet_names():
        std=workbook.get_sheet_by_name('Sheet')
        workbook.remove_sheet(std)
    workbook.save(os.path.join(results_path, name))
    # rows_groups = rg
    # not_used_cols = nuc

def Keyareas_columns(keyareas, sorting_lists, mandatory_word = None, complementary_info = None, computations = 'sum', forbiden_word = ''):
    if mandatory_word != None and complementary_info != None:
        keyword = mandatory_word+complementary_info
    elif mandatory_word != None:
        keyword = mandatory_word
    else:
        keyword = ''
    all_columns = keyareas.columns.tolist()
    #general_cols = ['Id', 'RealScoring']
    general_cols = copy.deepcopy(not_used_cols)
    # for row in rows_groups:
    #     if len(rows_groups[row][0]) == 0:
    #         general_cols.append(row)
    #     else:
    #         general_cols += rows_groups[row][0]
    first_group_columns = {'data': keyareas,
                            'first_columns': None,
                            'second_columns': None,
                            'third_columns': None,
                            'computations_columns': None,
                            'columns': None}
    third_columns_names = list(set(all_columns) - set(general_cols))
    # we pre-estimate the columns to use (based on the need of showing section data or not)
    if keyword != '':
        if forbiden_word != '':
            third_columns_names = [c.replace(keyword, '') for c in third_columns_names if keyword in c and forbiden_word not in c]
        else:
            third_columns_names = [c.replace(keyword, '') for c in third_columns_names if keyword in c]
    else:
        third_columns_names = [c for c in third_columns_names if 'Section' not in c and 'Compasses' not in c and 'Areas' not in c and 'Modulatory' not in c and 'ModComp' not in c] #all the possibilities without keywords

    # We pick only the columns that represent 'Key'
    cleaned_third_columns_names = [tcn.replace('Key', '') for tcn in third_columns_names if 'Key' in tcn and 'Groupping' not in tcn]

    third_columns_names = sort(cleaned_third_columns_names, sorting_lists['Modulation'])
    computations_list = [computations]*len(cleaned_third_columns_names) # This sheet is about adding, so only sum is computed
    first_group_columns['second_column_names'] = [('', 1), ('Key', len(cleaned_third_columns_names))]
    if keyword != '':
        column_names =  columns_alike_our_data(third_columns_names, [(keyword, len(third_columns_names))], [('Key', len(third_columns_names))])
    else:
        column_names =  columns_alike_our_data(third_columns_names, [('Key', len(third_columns_names))])
    first_group_columns['third_columns_names'] = ['Total analysed'] + third_columns_names
    first_group_columns['column_names'] = ['Total analysed'] + column_names
    first_group_columns['computations_columns'] = ['sum'] + computations_list
    
    # Creamos los datos para las segundas agrupaciones
    third_columns_names = list(set(all_columns) - set(general_cols))
    if keyword != '':
        if forbiden_word != '':
            third_columns_names = [c.replace(keyword, '') for c in third_columns_names if keyword in c and forbiden_word not in c]
        else:
            third_columns_names = [c.replace(keyword, '') for c in third_columns_names if keyword in c]
    else:
        third_columns_names = [c for c in third_columns_names if 'Section' not in c and 'Compasses' not in c and 'Areas' not in c and 'Modulatory' not in c and 'ModComp' not in c]

    second_group_columns = {'data': None,
                            'first_columns': None,
                            'second_columns': None,
                            'third_columns': None,
                            'computations_columns': None,
                            'columns': None}
    cleaned_third_columns_names = [tcn.replace('KeyGroupping1', '') for tcn in third_columns_names if 'KeyGroupping1' in tcn]
    second_group_columns['third_columns'] = sort(cleaned_third_columns_names, sorting_lists['ModulationG1Sorting'])
    second_group_columns['second_columns'] = [('', 1), ('KeyGroupping1', len(cleaned_third_columns_names))]
    second_group_columns['computations_columns'] = [computations]*len(cleaned_third_columns_names) # esta hoja va de sumar, así que en todas las columnas el cómputo que hay que hacer es sumar!
    if keyword != '':
        second_group_columns['columns'] = columns_alike_our_data(second_group_columns['third_columns'], [(keyword, len(cleaned_third_columns_names))], [('KeyGroupping1', len(cleaned_third_columns_names))])
    else:
        second_group_columns['columns'] = columns_alike_our_data(second_group_columns['third_columns'], [('KeyGroupping1', len(cleaned_third_columns_names))])

    second_group_columns['data'] = keyareas
    second_group_columns['third_columns'] = ['Total analysed'] + second_group_columns['third_columns']
    second_group_columns['computations_columns'] = ['sum'] + second_group_columns['computations_columns']
    second_group_columns['columns'] = ['Total analysed'] + second_group_columns['columns']
    
    # creamos los datos para las terceras agrupaciones
    third_columns_names = list(set(all_columns) - set(general_cols))
    if keyword != '':
        if forbiden_word != '':
            third_columns_names = [c.replace(keyword, '') for c in third_columns_names if keyword in c and forbiden_word not in c]
        else:
            third_columns_names = [c.replace(keyword, '') for c in third_columns_names if keyword in c]
    else:
        third_columns_names = [c for c in third_columns_names if 'Section' not in c and 'Compasses' not in c and 'Areas' not in c and 'Modulatory' not in c and 'ModComp' not in c]

    third_group_columns = {'data': None,
                            'first_columns': None,
                            'second_columns': None,
                            'third_columns': None,
                            'computations_columns': None,
                            'columns': None}
    cleaned_third_columns_names = [tcn.replace('KeyGroupping2', '') for tcn in third_columns_names if 'KeyGroupping2' in tcn]
    third_group_columns['third_columns'] = sort(cleaned_third_columns_names, sorting_lists['ModulationG2Sorting'])
    third_group_columns['second_columns'] = [('', 1), ('KeyGroupping2', len(cleaned_third_columns_names))]
    third_group_columns['computations_columns'] = [computations]*len(cleaned_third_columns_names) # esta hoja va de sumar, así que en todas las columnas el cómputo que hay que hacer es sumar!
    if keyword != '':
        third_group_columns['columns'] = columns_alike_our_data(third_group_columns['third_columns'], [(keyword, len(cleaned_third_columns_names))], [('KeyGroupping2', len(cleaned_third_columns_names))])
    else:
        third_group_columns['columns'] = columns_alike_our_data(third_group_columns['third_columns'], [('KeyGroupping2', len(cleaned_third_columns_names))])
    third_group_columns['data'] = keyareas
    third_group_columns['third_columns'] = ['Total analysed'] + third_group_columns['third_columns']
    third_group_columns['computations_columns'] = ['sum'] + third_group_columns['computations_columns']
    third_group_columns['columns'] = ['Total analysed'] + third_group_columns['columns']

    return first_group_columns, second_group_columns, third_group_columns

########################################################################
# Function in charge of generating Ia.keyareas
########################################################################
def Keyareas(results_path, keyareas, combinations = False, groups = None):
    # additional_info = {"Label":["Aria"], "Aria":['Label']}
    workbook = openpyxl.Workbook()
    first_group_columns,second_group_columns,third_group_columns = Keyareas_columns(keyareas)
    first_group_columns_A,second_group_columns_A,third_group_columns_A = Keyareas_columns(keyareas, mandatory_word = 'Section', complementary_info = 'A', forbiden_word='ModComp')
    first_group_columns_B,second_group_columns_B,third_group_columns_B = Keyareas_columns(keyareas, mandatory_word = 'Section', complementary_info = 'B', forbiden_word='ModComp')
    
    print('\t\t Horizontal per')
    excel_sheet(workbook.create_sheet("Horizontal Per"), first_group_columns['column_names'], keyareas, first_group_columns['third_columns_names'], first_group_columns['computations_columns'], 
                second_columns=first_group_columns['second_column_names'], per = True, average=True, last_column=True, last_column_average = False, additional_info=additional_info,
                second_subgroup=True, second_subgroup_info=second_group_columns, 
                third_subgroup=True, third_subgroup_info=third_group_columns,
                groups = groups)
    print('\t\t Vertical per')
    excel_sheet(workbook.create_sheet("Vertical Per"), first_group_columns['column_names'], keyareas, first_group_columns['third_columns_names'], first_group_columns['computations_columns'], 
                second_columns=first_group_columns['second_column_names'], per = True, average = False, last_column=True, last_column_average = True, additional_info=additional_info,
                second_subgroup=True, second_subgroup_info=second_group_columns,
                third_subgroup=True, third_subgroup_info=third_group_columns,
                groups = groups)
    print('\t\t Horizontal per A')
    excel_sheet(workbook.create_sheet("Horizontal Per A"), first_group_columns_A['column_names'], keyareas, first_group_columns_A['third_columns_names'], first_group_columns_A['computations_columns'], 
                second_columns=first_group_columns_A['second_column_names'], per = True, average=True, last_column=True, last_column_average = False, additional_info=additional_info,
                second_subgroup=True, second_subgroup_info=second_group_columns_A, 
                third_subgroup=True, third_subgroup_info=third_group_columns_A,
                groups = groups)
    print('\t\t Vertical per A')
    excel_sheet(workbook.create_sheet("Vertical Per A"), first_group_columns_A['column_names'], keyareas, first_group_columns_A['third_columns_names'], first_group_columns_A['computations_columns'], 
                second_columns=first_group_columns_A['second_column_names'], per = True, average = False, last_column=True, last_column_average = True, additional_info=additional_info,
                second_subgroup=True, second_subgroup_info=second_group_columns_A,
                third_subgroup=True, third_subgroup_info=third_group_columns_A,
                groups = groups)
    print('\t\t Horizontal per B')
    excel_sheet(workbook.create_sheet("Horizontal Per B"), first_group_columns_B['column_names'], keyareas, first_group_columns_B['third_columns_names'], first_group_columns_B['computations_columns'], 
                second_columns=first_group_columns_B['second_column_names'], per = True, average=True, last_column=True, last_column_average = False, additional_info=additional_info,
                second_subgroup=True, second_subgroup_info=second_group_columns_B, 
                third_subgroup=True, third_subgroup_info=third_group_columns_B,
                groups = groups)
    print('\t\t Horizontal per B')
    excel_sheet(workbook.create_sheet("Vertical Per B"), first_group_columns_B['column_names'], keyareas, first_group_columns_B['third_columns_names'], first_group_columns_B['computations_columns'], 
                second_columns=first_group_columns_B['second_column_names'], per = True, average = False, last_column=True, last_column_average = True, additional_info=additional_info,
                second_subgroup=True, second_subgroup_info=second_group_columns_B,
                third_subgroup=True, third_subgroup_info=third_group_columns_B,
                groups = groups)
    #borramos la hoja por defecto
    if "Sheet" in workbook.get_sheet_names():
        std=workbook.get_sheet_by_name('Sheet')
        workbook.remove_sheet(std)
    workbook.save(os.path.join(results_path, "Ia.Keyareas.xlsx"))
########################################################################
# Function in charge of generating Ib.Keyareas_weighted
########################################################################

# def Keyareas_weighted(results_path, keyareas, combinations = False, groups = None):
def Keyareas_weighted(_cfg: Configuration, data: DataFrame, results_path: str, name: str, sorting_lists: list, visualiser_lock: Lock, groups: list=None, additional_info: list=[]):

    # additional_info = {"Label":["Aria"], "Aria":['Label']}
    workbook = openpyxl.Workbook()
    ############
    # COMPASES #
    ############
    first_group_columns,second_group_columns,third_group_columns = Keyareas_columns(data,sorting_lists, 'Compasses', computations = 'mean')
    
    excel_sheet(workbook.create_sheet("WH_proportions"), first_group_columns['column_names'], data, first_group_columns['third_columns_names'], first_group_columns['computations_columns'], 
                second_columns=first_group_columns['second_column_names'], per = True, average=True, last_column=True, last_column_average = False, additional_info=additional_info,
                
                
                second_subgroup=True, second_subgroup_info=second_group_columns, 
                third_subgroup=True, third_subgroup_info=third_group_columns,
                groups = groups)
                
    excel_sheet(workbook.create_sheet("WV_proportions"), first_group_columns['column_names'], data, first_group_columns['third_columns_names'], first_group_columns['computations_columns'], 
                second_columns=first_group_columns['second_column_names'], per = True, average = False, last_column=True, last_column_average = True, additional_info=additional_info,
                second_subgroup=True, second_subgroup_info=second_group_columns,
                third_subgroup=True, third_subgroup_info=third_group_columns,
                groups = groups)
    ############
    # AREAS    #
    ############
    first_group_columns,second_group_columns,third_group_columns = Keyareas_columns(data, sorting_lists, 'Modulatory', computations = 'mean')
    excel_sheet(workbook.create_sheet("WH_key_areas"), first_group_columns['column_names'], data, first_group_columns['third_columns_names'], first_group_columns['computations_columns'], 
                second_columns=first_group_columns['second_column_names'], per = True, average=True, last_column=True, last_column_average = False, additional_info=additional_info,
                second_subgroup=True, second_subgroup_info=second_group_columns, 
                third_subgroup=True, third_subgroup_info=third_group_columns,
                groups = groups) #This is the main case in which the weighted FLAG is turned on
    #################
    # ALLWEIGHTED   #
    #################
    first_group_columns,second_group_columns,third_group_columns = Keyareas_columns(data, sorting_lists, 'ModComp', computations = 'mean', forbiden_word='Section')
    excel_sheet(workbook.create_sheet("WH_allweighted"), first_group_columns['column_names'], data, first_group_columns['third_columns_names'], first_group_columns['computations_columns'], 
                second_columns=first_group_columns['second_column_names'], per = True, average=True, last_column=True, last_column_average = False, additional_info=additional_info,
                second_subgroup=True, second_subgroup_info=second_group_columns, 
                third_subgroup=True, third_subgroup_info=third_group_columns,
                groups = groups) #This is the main case in which the weighted FLAG is turned on
    excel_sheet(workbook.create_sheet("WV_allweighted"), first_group_columns['column_names'], data, first_group_columns['third_columns_names'], first_group_columns['computations_columns'], 
                second_columns=first_group_columns['second_column_names'], per = True, average=False, last_column=True, last_column_average = True, additional_info=additional_info,
                second_subgroup=True, second_subgroup_info=second_group_columns, 
                third_subgroup=True, third_subgroup_info=third_group_columns,
                groups = groups) #This is the main case in which the weighted FLAG is turned on
    
    first_group_columns_A,second_group_columns_A,third_group_columns_A = Keyareas_columns(data, sorting_lists, mandatory_word = 'ModCompSection', complementary_info = 'A')
    first_group_columns_B,second_group_columns_B,third_group_columns_B = Keyareas_columns(data, sorting_lists, mandatory_word = 'ModCompSection', complementary_info = 'B')
    #################
    # SECTION A   #
    #################
    excel_sheet(workbook.create_sheet("WH_allweighted_A"), first_group_columns_A['column_names'], data, first_group_columns_A['third_columns_names'], first_group_columns_A['computations_columns'], 
                second_columns=first_group_columns_A['second_column_names'], per = True, average=True, last_column=True, last_column_average = False, additional_info=additional_info,
                second_subgroup=True, second_subgroup_info=second_group_columns_A, 
                third_subgroup=True, third_subgroup_info=third_group_columns_A,
                groups = groups) #This is the main case in which the weighted FLAG is turned on
    excel_sheet(workbook.create_sheet("WV_allweighted_A"), first_group_columns_A['column_names'], data, first_group_columns_A['third_columns_names'], first_group_columns_A['computations_columns'], 
                second_columns=first_group_columns_A['second_column_names'], per = True, average=False, last_column=True, last_column_average = True, additional_info=additional_info,
                second_subgroup=True, second_subgroup_info=second_group_columns_A, 
                third_subgroup=True, third_subgroup_info=second_group_columns_A,
                groups = groups) #This is the main case in which the weighted FLAG is turned on
    #################
    # SECTION B   #
    #################
    excel_sheet(workbook.create_sheet("WH_allweighted_B"), first_group_columns_B['column_names'], data, first_group_columns_B['third_columns_names'], first_group_columns_B['computations_columns'], 
                second_columns=first_group_columns_B['second_column_names'], per = True, average=True, last_column=True, last_column_average = False, additional_info=additional_info,
                second_subgroup=True, second_subgroup_info=second_group_columns_B, 
                third_subgroup=True, third_subgroup_info=third_group_columns_B,
                groups = groups) #This is the main case in which the weighted FLAG is turned on
    excel_sheet(workbook.create_sheet("WV_allweighted_B"), first_group_columns_B['column_names'], data, first_group_columns_B['third_columns_names'], first_group_columns_B['computations_columns'], 
                second_columns=first_group_columns_B['second_column_names'], per = True, average=False, last_column=True, last_column_average = True, additional_info=additional_info,
                second_subgroup=True, second_subgroup_info=second_group_columns_B, 
                third_subgroup=True, third_subgroup_info=third_group_columns_B,
                groups = groups) #This is the main case in which the weighted FLAG is turned on

    if "Sheet" in workbook.get_sheet_names():
        std=workbook.get_sheet_by_name('Sheet')
        workbook.remove_sheet(std)
    workbook.save(os.path.join(results_path, "Ib.Keyareas_weighted.xlsx"))

def Chords(factor, _cfg: Configuration, data: DataFrame, results_path: str, name: str, visualiser_lock: Lock, groups: list=None, additional_info: list=[]):
    try:
        workbook = openpyxl.Workbook()
        # Splitting the dataframes to reorder them
        data_general = data[metadata_columns + ['Total analysed']].copy()
        data = data[set(data.columns).difference(metadata_columns)]
        data_general = data_general.dropna(how='all', axis=1)
        data = data.dropna(how='all', axis=1)
        data.columns = [i.replace('_','') for i in data.columns]
        data.columns=[i.replace('Chords', '') for i in data.columns]
        
        cols = sort(data.columns.tolist(), [
                    i for i in _cfg.sorting_lists['NumeralsSorting']])
        
        data = data[cols]
        third_columns_names = data.columns.to_list()

        second_column_names = [("", 1), ("Chords", len(third_columns_names))]
        third_columns_names.insert(0, 'Total analysed')
        data = pd.concat([data_general, data], axis=1)
        # data_total = pd.concat([data_general, notes_and_measures], axis=1)

        computations = ["sum"] + ["mean"] * (len(third_columns_names)-1)
        # computations2 = ["sum"] + ["mean_density"] * \
        #     (len(third_columns_names)-1)
        columns = third_columns_names
        excel_sheet(workbook.create_sheet("Weighted"), columns, data, third_columns_names, computations, _cfg.sorting_lists, groups=groups, last_column=True,
                     last_column_average=True, second_columns=second_column_names, average=True, additional_info=additional_info, ponderate=False)
        # if factor >=1:
        #     excel_sheet(workbook.create_sheet("Horizontal"), columns, data_total, third_columns_names, computations2,  _cfg.sorting_lists, groups=groups,
        #              second_columns=second_column_names, per=False, average=True, last_column=True, last_column_average=True, additional_info=additional_info)

        # Deleting default sheet
        if "Sheet" in workbook.get_sheet_names():
            std = workbook.get_sheet_by_name('Sheet')
            workbook.remove_sheet(std)
        adjust_excel_width_height(workbook)
        workbook.save(os.path.join(results_path, name))

        # with visualiser_lock: #Apply when threads are usedwith visualizer_lock=threading.Lock()
        columns.remove('Total analysed')
        title = 'Instrumental densities'
        # VISUALISATIONS
        # if groups:
        #     data_grouped = data.groupby(list(groups))
        #     for g, datag in data_grouped:
        #         result_visualisations = results_path + \
        #             '\\visualisations\\' + str(g.replace('/', '_'))
        #         if not os.path.exists(result_visualisations):
        #             os.mkdir(result_visualisations)
        #         name_bar = result_visualisations + \
        #             '\\' + name.replace('.xlsx', '.png')
        #         bar_plot_extended(name_bar, datag, columns, 'Density',
        #                           'Density', title, second_title=str(g))

        # elif factor == 1:  # 
        #     groups = [i for i in rows_groups]
        #     for row in rows_groups:
        #         plot_name = name.replace(
        #             '.xlsx', '') + '_Per_' + str(row.upper()) + '.png'
        #         name_bar = results_path + '\\visualisations\\' + plot_name
        #         if row not in not_used_cols:
        #             if len(rows_groups[row][0]) == 0:  # no sub-groups
        #                 data_grouped = data.groupby(row, sort=True)
        #                 if data_grouped:
        #                     line_plot_extended(
        #                     name_bar, data_grouped, columns, 'Instrument', 'Density', title, second_title='Per ' + str(row))
        #             else:
        #                 for i, subrow in enumerate(rows_groups[row][0]):
        #                     if subrow not in EXCEPTIONS:
        #                         plot_name = name.replace(
        #                             '.xlsx', '') + '_Per_' + str(row.upper()) + '_' + str(subrow) + '.png'
        #                         name_bar = results_path + '\\visualisations\\' + plot_name
        #                         data_grouped = data.groupby(subrow)
        #                         line_plot_extended(
        #                             name_bar, data_grouped, columns, 'Instrument', 'Density', title, second_title= 'Per ' + str(subrow))
        # else:
        #     for instr in third_columns_names:
        #         columns=data['AriaName']
        #         name_bar = results_path + '\\visualisations\\' + instr + \
        #             name.replace('.xlsx', '.png')
        #         bar_plot_extended(name_bar, data, columns,
        #                         'Density', 'Density', title + ' ' + instr, instr=instr)
    except Exception as e:
        _cfg.write_logger.error('{}  Problem found: {}'.format(name, e))
# def Chords(chords, results_path, name, sorting, keyword=''):
#     additional_info = {"Label":["Aria"], "Aria":['Label']}
#     workbook = openpyxl.Workbook()
    
#     all_columns = chords.columns.tolist()
#     #general_cols = ['Id', 'RealScoring']
#     general_cols = copy.deepcopy(not_used_cols)
#     for row in rows_groups:
#         if len(rows_groups[row][0]) == 0:
#             general_cols.append(row)
#         else:
#             general_cols += rows_groups[row][0]
#     first_group_columns = {'data': chords,
#                             'first_columns': None,
#                             'second_columns': None,
#                             'third_columns': None,
#                             'computations_columns': None,
#                             'columns': None}
#     third_columns_names = list(set(all_columns) - set(general_cols))
#     # we pre-estimate the columns to use (based on the need of showing section data or not)
#     if keyword != '':
#         third_columns_names = [c.replace(keyword, '') for c in third_columns_names if keyword in c]
#     else:
#         third_columns_names = [c for c in third_columns_names if 'Groupping' not in c] #all the possibilities without keywords
#     cleaned_third_columns_names = [tcn.replace('Chords', '').strip() for tcn in third_columns_names if 'Chord' in tcn]
#     #third_columns_names = sort(cleaned_third_columns_names, sorting_lists[sorting])
#     relative_root_sorting = harmony_sorting.get_relativeroot_sorting()
#     numerals_sorting = harmony_sorting.get_degrees_sorting()
#     form_sorting = harmony_sorting.get_chordform_sorting()
#     figbass_sorting = harmony_sorting.get_inversions_sorting()
#     if sorting == "ChordSorting":
#         third_columns_names = list(lausanne_sorting_script.sort_labels(cleaned_third_columns_names, relativeroot=relative_root_sorting, numeral=numerals_sorting, form=form_sorting, figbass=figbass_sorting, changes='occurrences'))
#     else:
#         third_columns_names = sort(cleaned_third_columns_names, sorting_lists[sorting])

#     computations_list = ['sum']*len(cleaned_third_columns_names) # esta hoja va de sumar, así que en todas las columnas el cómputo que hay que hacer es sumar!
#     first_group_columns['second_column_names'] = [('', 1), ('Chords', len(cleaned_third_columns_names))]
#     if keyword != '':
#         column_names =  columns_alike_our_data(third_columns_names, [(keyword, len(third_columns_names))], [('Chords', len(third_columns_names))])
#     else:
#         column_names =  columns_alike_our_data(third_columns_names, [('Chords', len(third_columns_names))])
#     first_group_columns['third_columns_names'] = ['Total analysed'] + third_columns_names
#     first_group_columns['column_names'] = ['Total analysed'] + column_names
#     first_group_columns['computations_columns'] = ['sum'] + computations_list

#     print('\t\t H_total')
#     excel_sheet(workbook.create_sheet("H_total"), first_group_columns['column_names'], chords, first_group_columns['third_columns_names'], first_group_columns['computations_columns'], 
#                 second_columns=first_group_columns['second_column_names'], per = True, average=True, last_column=True, last_column_average = False, additional_info=additional_info)

#     #borramos la hoja por defecto
#     if "Sheet" in workbook.get_sheet_names():
#         std=workbook.get_sheet_by_name('Sheet')
#         workbook.remove_sheet(std)
#     workbook.save(os.path.join(results_path, name))


def get_sorting_lists():
    RoleSorting = general_sorting.get_role_sorting() # Only valid for DIDONE corpus, any other roles will be sorted alphabetically
    FormSorting = general_sorting.get_form_sorting()
    KeySorting = general_sorting.get_key_sorting()
    KeySignatureSorting = general_sorting.get_KeySignature_sorting()
    KeySignatureGroupedSorted = general_sorting.get_KeySignatureType_sorting()
    TimeSignatureSorting = general_sorting.get_TimeSignature_sorting()
    TempoSorting = general_sorting.get_Tempo_sorting()
    TempoGroupedSorting1 = general_sorting.get_TempoGrouped1_sorting()
    TempoGroupedSorting2 = general_sorting.get_TempoGrouped2_sorting()
    clefs = general_sorting.get_Clefs_sorting()
    scoring_sorting = general_sorting.get_scoring_sorting() #Long combination
    scoring_family = general_sorting.get_familiesCombinations_sorting()
    # lista de intervalos (columnas en los archivos 2 y 3)
    Intervals = melody_sorting.intervals_sorting()
    Intervals_absolute = melody_sorting.intervals_absolutte_sorting()
    scale_degrees = melody_sorting.MelodicDegrees_sorting()
    #harmony
    modulations = harmony_sorting.get_modulations_sorting()
    modulationsG1 = harmony_sorting.get_modulationsGrouping1_sorting()
    modulationsG2 = harmony_sorting.get_modulationsGrouping2_sorting()
    chordTypes, chordTypesG = harmony_sorting.get_chordtype_sorting()
    return {"RoleSorting": [i.lower() for i in RoleSorting],
            "FormSorting": [i.lower() for i in FormSorting],
            "KeySorting": KeySorting,
            "KeySignatureSorting": KeySignatureSorting,
            "KeySignatureGroupedSorted": KeySignatureGroupedSorted,
            "TimeSignatureSorting": TimeSignatureSorting,
            "TempoSorting": TempoSorting + [''], #a veces algunas puede ser nan, ya que no tienen tempo mark, las nan las ponemos al final
            "TempoGroupedSorting1": TempoGroupedSorting1 + [''],
            "TempoGroupedSorting2": TempoGroupedSorting2 + [''],
            "Intervals": Intervals,
            "Intervals_absolute": Intervals_absolute,
            "Clefs": clefs,
            "ScoringSorting": scoring_sorting,
            "ScoringFamilySorting": scoring_family,
            "ScaleDegrees": scale_degrees,
            "ModulationG1Sorting": modulationsG1,
            "ModulationG2Sorting": modulationsG2,
            "Modulation": modulations,
            "ChordTypeGrouppingSorting": chordTypesG,
            }


